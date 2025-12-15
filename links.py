import os
import re
import shutil
import tempfile
import pandas as pd
import streamlit as st
import pytesseract
import en_core_web_sm
import fitz
from PyPDF2 import PdfReader
from concurrent.futures import ThreadPoolExecutor, as_completed
import asyncio
import aiohttp
import google.generativeai as genai
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin
import atexit


# =====================================================
# SAFE TEMP DIRECTORY (HUGGINGFACE)
# =====================================================
DOWNLOAD_PATH = "/tmp/downloads"
os.makedirs(DOWNLOAD_PATH, exist_ok=True)

if "temp_files" not in st.session_state:
    st.session_state.temp_files = []


# =====================================================
# CLEANUP AFTER SESSION ENDS
# =====================================================
def full_cleanup():
    try:
        if os.path.exists(DOWNLOAD_PATH):
            shutil.rmtree(DOWNLOAD_PATH, ignore_errors=True)
        os.makedirs(DOWNLOAD_PATH, exist_ok=True)

        for f in st.session_state.get("temp_files", []):
            if os.path.exists(f):
                os.remove(f)

        st.session_state.temp_files = []
    except:
        pass


try:
    st.on_session_end(full_cleanup)
except:
    atexit.register(full_cleanup)


# =====================================================
# INITIALIZE OCR + NLP
# =====================================================
pytesseract.pytesseract.tesseract_cmd = "tesseract"
nlp = en_core_web_sm.load()

EMAIL_REGEX = r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}"

API_KEY = os.getenv("GENAI_KEY")
if API_KEY:
    genai.configure(api_key=API_KEY)
    gemini = genai.GenerativeModel("gemini-1.5-flash")


# =====================================================
# UNIVERSAL PDF LINK EXTRACTOR
# =====================================================
def extract_pdf_links(issue_url):
    headers = {"User-Agent": "Mozilla/5.0"}
    resp = requests.get(issue_url, headers=headers, timeout=30)
    resp.raise_for_status()
    html = resp.text
    soup = BeautifulSoup(html, "lxml")

    pdfs = set()
    pdf_regex = r'https?:\/\/[^\s"\']+\.pdf'

    # 1) Direct PDF hrefs
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if ".pdf" in href.lower():
            pdfs.add(urljoin(issue_url, href))

    # 2) OJS galleys
    for a in soup.select("a.obj_galley_link, a.galley-link, a.pdf"):
        href = a.get("href")
        if href:
            full = urljoin(issue_url, href)
            if "/view/" in full:
                full = full.replace("/view/", "/download/")
            pdfs.add(full)

    # 3) Text-based PDF buttons
    for a in soup.find_all("a", string=True, href=True):
        txt = a.get_text(strip=True).lower()
        if "pdf" in txt or "download" in txt:
            pdfs.add(urljoin(issue_url, a["href"]))

    # 4) Links inside <script> tags
    for script in soup.find_all("script"):
        if script.string:
            found = re.findall(pdf_regex, script.string)
            for link in found:
                pdfs.add(link)

    # 5) JSON-LD metadata
    for script in soup.find_all("script", type="application/ld+json"):
        if script.string:
            found = re.findall(pdf_regex, script.string)
            for link in found:
                pdfs.add(link)

    # 6) Fallback: search entire HTML
    for link in re.findall(pdf_regex, html):
        pdfs.add(link)

    return list(pdfs)


# =====================================================
# ASYNC AIOHTTP DOWNLOADER (MAX SPEED)
# =====================================================
async def fetch_pdf(session, link, idx):
    filename = os.path.join(DOWNLOAD_PATH, f"article_{idx}.pdf")

    try:
        async with session.get(link, timeout=60) as resp:
            if resp.status != 200:
                return (link, False, f"HTTP {resp.status}")

            ctype = resp.headers.get("Content-Type", "").lower()
            if "pdf" not in ctype:
                return (link, False, "Invalid PDF")

            with open(filename, "wb") as f:
                async for chunk in resp.content.iter_chunked(1024 * 256):
                    f.write(chunk)

            st.session_state.temp_files.append(filename)
            return (filename, True, "Downloaded")

    except Exception as e:
        return (link, False, str(e))


async def download_all_async(pdf_links):
    results = []
    total = len(pdf_links)

    progress = st.progress(0)
    done = 0

    connector = aiohttp.TCPConnector(limit=20)
    timeout = aiohttp.ClientTimeout(total=150)

    async with aiohttp.ClientSession(connector=connector, timeout=timeout) as session:
        tasks = [
            asyncio.create_task(fetch_pdf(session, link, i+1))
            for i, link in enumerate(pdf_links)
        ]

        for task in asyncio.as_completed(tasks):
            result = await task
            results.append(result)
            done += 1
            progress.progress(done / total)

    return results


def download_pdfs(pdf_links):
    return asyncio.run(download_all_async(pdf_links))


# =====================================================
# OCR TITLE + EMAIL EXTRACTOR
# =====================================================
def extract_title_fast(pdf_path):
    blacklist = ["journal", "springer", "elsevier", "wiley", "sage", "nature"]

    try:
        doc = fitz.open(pdf_path)
        page = doc[0]
    except:
        return "Unknown Title"

    spans = []
    for block in page.get_text("dict")["blocks"]:
        if "lines" not in block:
            continue

        for ln in block["lines"]:
            text = " ".join(s["text"] for s in ln["spans"]).strip()
            if not text:
                continue
            size = max(s["size"] for s in ln["spans"])
            spans.append((size, text))

    if not spans:
        return "Unknown Title"

    spans.sort(reverse=True)
    top_size = spans[0][0]

    candidates = [t for s, t in spans if abs(s - top_size) < 1.2]
    filtered = [t for t in candidates if not any(b in t.lower() for b in blacklist)]

    title = filtered[0] if filtered else spans[0][1]
    return title if len(title.split()) >= 4 else "Unknown Title"


def extract_text_fast(pdf_path):
    try:
        reader = PdfReader(pdf_path)
        t = reader.pages[0].extract_text()
        if t:
            return t
    except:
        pass

    try:
        doc = fitz.open(pdf_path)
        t = doc[0].get_text()
        if t:
            return t
    except:
        pass

    try:
        doc = fitz.open(pdf_path)
        pix = doc[0].get_pixmap()
        return pytesseract.image_to_string(pix.tobytes("png"))
    except:
        return ""


def extract_info_smart(text):
    emails = list(set(re.findall(EMAIL_REGEX, text)))
    persons = [ent.text for ent in nlp(text).ents if ent.label_ == "PERSON"]

    pairs = []
    for email in emails:
        epos = text.find(email)
        best, closest = 999999, ""

        for name in persons:
            npos = text.find(name)
            if npos != -1:
                d = abs(epos - npos)
                if d < best:
                    best = d
                    closest = name

        pairs.append((closest, email))

    return [p[0] for p in pairs], [p[1] for p in pairs]


def process_pdf(file):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        tmp.write(file.read())
        path = tmp.name

    st.session_state.temp_files.append(path)

    title = extract_title_fast(path)
    text = extract_text_fast(path)
    names, emails = extract_info_smart(text)

    return file.name, title, names, emails


# =====================================================
# EXPORT TO EXCEL
# =====================================================
def create_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    for col, name in enumerate(df.columns, 1):
        c = ws.cell(row=1, column=col, value=name)
        c.font = header_font
        c.fill = header_fill
        c.border = border

    for r in range(len(df)):
        for c in range(len(df.columns)):
            ws.cell(row=r+2, column=c+1, value=str(df.iloc[r, c])).border = border

    path = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
    wb.save(path)
    st.session_state.temp_files.append(path)

    return path


# =====================================================
# STREAMLIT UI
# =====================================================
st.set_page_config(page_title="PDF Toolkit", layout="wide")
st.title("📚 Ultra-Fast PDF Downloader + OCR Email Extractor")


tab1, tab2 = st.tabs(["📥 PDF Downloader", "📧 OCR Email Extractor"])


# ============================== TAB 1 ==============================
with tab1:
    st.subheader("Download PDFs from Journal Issue URL")

    url = st.text_input("Enter Issue URL:")

    if st.button("Extract PDF Links"):
        links = extract_pdf_links(url)
        st.session_state.links = links

        st.success(f"Found {len(links)} PDF link(s).")
        for link in links:
            st.write(link)

    if st.session_state.get("links"):
        if st.button("Download All PDFs (Ultra-Fast Async)"):
            st.warning("Downloading PDFs… please wait.")
            results = download_pdfs(st.session_state.links)

            st.write("### Download Report:")
            for file, ok, msg in results:
                if ok:
                    st.success(file)
                else:
                    st.error(f"{file} → {msg}")

            # Create ZIP
            zip_base = "/tmp/all_pdfs"
            zip_path = f"{zip_base}.zip"

            shutil.make_archive(zip_base, "zip", DOWNLOAD_PATH)
            st.session_state.temp_files.append(zip_path)

            with open(zip_path, "rb") as f:
                st.download_button(
                    label="⬇️ Download All PDFs (ZIP)",
                    data=f.read(),
                    file_name="all_pdfs.zip",
                    mime="application/zip",
                )

            st.info("📦 ZIP ready! Files will auto-delete once your session ends.")


# ============================== TAB 2 ==============================
with tab2:
    st.subheader("Extract Titles, Authors & Emails (OCR Enabled)")

    uploaded = st.file_uploader("Upload PDFs (max 500)", type=["pdf"], accept_multiple_files=True)

    if uploaded:
        total = len(uploaded)

        if total > 500:
            st.error("❌ Maximum 500 PDFs allowed.")
            st.stop()

        st.info(f"Processing {total} PDFs…")

        rows = []
        progress = st.progress(0)
        done = 0

        with ThreadPoolExecutor(max_workers=8) as exe:
            futures = {exe.submit(process_pdf, f): f for f in uploaded}

            for future in as_completed(futures):
                fname, title, names, emails = future.result()

                for n, e in zip(names, emails):
                    rows.append([fname, title, n, e])

                done += 1
                progress.progress(done / total)

        df = pd.DataFrame(rows, columns=["PDF File", "Title", "Author", "Email"])
        st.dataframe(df, use_container_width=True)

        st.download_button("⬇️ Download CSV", df.to_csv(index=False), "emails.csv")

        xfile = create_excel(df)
        st.download_button("⬇️ Download Excel", open(xfile, "rb"), "emails.xlsx")